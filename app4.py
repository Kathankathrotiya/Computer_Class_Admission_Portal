# app.py
from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from flask_cors import CORS
app = Flask(__name__)
CORS(app)

def ensure_excel_files():
    ensure_student_file()
    ensure_batch_file()
    ensure_completion_file()

def ensure_student_file():
    file_path = 'student_data.xlsx'
    if not Path(file_path).is_file():
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['ID', 'Name', 'Address', 'City', 'Mobile No 1', 'Mobile No 2','Standard Studying','School','Date of Birth','Father Occupation','Course Done','Which','Where', 'Course Name','Starting Date', 'Batch','Fees', 'Discount', 'Final Fees'])
        workbook.save(file_path)

def ensure_batch_file():
    file_path = 'batch_data.xlsx'
    if not Path(file_path).is_file():
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

def ensure_inquiry_file():
    file_path = 'inquiry_data.xlsx'
    if not Path(file_path).is_file():
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['ID', 'Inquiry Date','Name','City', 'Mobile No','Course Name','Batch','Starting Date'])
        workbook.save(file_path)

def ensure_completion_file():
    file_path = 'completion_data.xlsx'
    if not Path(file_path).is_file():
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['ID', 'Name', 'Batch','Course','Mobile No','Completion Date', 'Exam Date', 'Certificate Number', 'Issue Certificate Date', 'Receiver Name', 'Final Fees'])
        workbook.save(file_path)

def save_to_excel(data):
    ensure_excel_files()

    student_workbook = load_workbook('student_data.xlsx')
    student_sheet = student_workbook['Sheet']

    # Find the first empty row
    row_index = 2
    while student_sheet.cell(row=row_index, column=1).value is not None:
        row_index += 1

    # Use the same ID for both student and batch data
    student_id = row_index - 1

    student_workbook.close();
    student_workbook.save('student_data.xlsx')

    # Save data to batch_data.xlsx
    batch_workbook = load_workbook('batch_data.xlsx')

    batch_name = data['batch']
    if batch_name not in batch_workbook.sheetnames:
        batch_workbook.create_sheet(title=batch_name)

    batch_sheet = batch_workbook[batch_name]

    # Add header row if sheet is newly created
    if batch_sheet.max_row == 1:
        batch_sheet.append(['ID', 'Name', 'Mobile No 1', 'City','Course','PC Number','','Final Fees','Installment_1_Amount','Installment_1_Date','Installment_2_Amount','Installment_2_Date','Installment_3_Amount','Installment_3_Date','Fees Remaining','Completion Date'])

    
    # Find the first empty row
    row_index = batch_sheet.max_row + 1

    if row_index >= 22:
        return "Batch is full. Maximum admissions reached."

    # Set the initial remaining fees to the final fees submitted in the form
    initial_remaining_fees = data['finalFees']

    # Update the batch data Excel file with the new data
    batch_sheet.cell(row=row_index, column=1, value=student_id)
    batch_sheet.cell(row=row_index, column=2, value=data['name'])
    batch_sheet.cell(row=row_index, column=3, value=data['moNo1'])
    batch_sheet.cell(row=row_index, column=4, value=data['city'])
    batch_sheet.cell(row=row_index, column=5, value=data['course'])
    batch_sheet.cell(row=row_index, column=6, value='') #PC
    batch_sheet.cell(row=row_index, column=8, value=data['finalFees']) #Final Fees
    batch_sheet.cell(row=row_index, column=9, value='') #1
    batch_sheet.cell(row=row_index, column=10, value='') #1
    batch_sheet.cell(row=row_index, column=11, value='') #2
    batch_sheet.cell(row=row_index, column=12, value='') #2
    batch_sheet.cell(row=row_index, column=13, value='') #3
    batch_sheet.cell(row=row_index, column=14, value='') #3
    batch_sheet.cell(row=row_index, column=15, value=initial_remaining_fees)
    batch_sheet.cell(row=row_index, column=16, value='') #Completion date

    batch_workbook.save('batch_data.xlsx')

    # Save data to student_data.xlsx

    student_workbook = load_workbook('student_data.xlsx')
    student_sheet = student_workbook['Sheet']

    # Find the first empty row
    row_index = 2
    while student_sheet.cell(row=row_index, column=1).value is not None:
        row_index += 1

    # Use the same ID for both student and batch data
    student_id = row_index - 1

    # Update the student data Excel file with the new data
    student_sheet.cell(row=row_index, column=1, value=student_id)
    student_sheet.cell(row=row_index, column=2, value=data['name'])
    student_sheet.cell(row=row_index, column=3, value=data['address'])
    student_sheet.cell(row=row_index, column=4, value=data['city'])
    student_sheet.cell(row=row_index, column=5, value=data['moNo1'])
    student_sheet.cell(row=row_index, column=6, value=data['moNo2'])
    student_sheet.cell(row=row_index, column=7, value=data['standard'])
    student_sheet.cell(row=row_index, column=8, value=data['school'])
    student_sheet.cell(row=row_index, column=9, value=data['dob'])
    student_sheet.cell(row=row_index, column=10, value=data['occ'])
    student_sheet.cell(row=row_index, column=11, value=data['course1'])
    student_sheet.cell(row=row_index, column=12, value=data['courseWhich'])
    student_sheet.cell(row=row_index, column=13, value=data['courseWhere'])
    student_sheet.cell(row=row_index, column=14, value=data['course'])
    student_sheet.cell(row=row_index, column=15, value=data['startDate'])
    student_sheet.cell(row=row_index, column=16, value=data['batch'])
    student_sheet.cell(row=row_index, column=17, value=data['fees'])
    student_sheet.cell(row=row_index, column=18, value=data['discount'])
    student_sheet.cell(row=row_index, column=19, value=data['finalFees'])

    student_workbook.save('student_data.xlsx')

    return "Admission Successful"

@app.route('/update_remaining_data', methods=['POST'])
def update_remaining_data():
    try:
        student_id = request.form.get('studentId')
        batch_name = request.form.get('batch')

        print(f"Received request to update data for Student ID {student_id} in Batch {batch_name}")

        batch_workbook = load_workbook('batch_data.xlsx')

        if batch_name in batch_workbook.sheetnames:
            batch_sheet = batch_workbook[batch_name]

            col_index = None
            for cell in batch_sheet[1]:
                if cell.value == 'ID':
                    col_index = cell.column

            if col_index is not None:
                row_index = None
                for row in batch_sheet.iter_rows(min_row=2, max_col=1, max_row=batch_sheet.max_row):
                    if row[0].value == int(student_id):
                        row_index = row[0].row
                        break

                if row_index is not None:
                    print(f"Found Student ID {student_id} in Batch {batch_name} at Row {row_index}")

                    # Update remaining data based on the submitted form data
                    batch_sheet.cell(row=row_index, column=6).value = request.form.get('pcNumber', '')  # Ensure pcNumber is correctly handled
                    batch_sheet.cell(row=row_index, column=9).value = request.form.get('installmentAmount1')
                    batch_sheet.cell(row=row_index, column=10).value = request.form.get('installmentDate1')
                    batch_sheet.cell(row=row_index, column=11).value = request.form.get('installmentAmount2')
                    batch_sheet.cell(row=row_index, column=12).value = request.form.get('installmentDate2')
                    batch_sheet.cell(row=row_index, column=13).value = request.form.get('installmentAmount3')
                    batch_sheet.cell(row=row_index, column=14).value = request.form.get('installmentDate3')
                    batch_sheet.cell(row=row_index, column=15).value = request.form.get('amounttobePaid')
                    batch_sheet.cell(row=row_index, column=16).value = request.form.get('completionDate')

                    batch_workbook.save('batch_data.xlsx')

                    # Check if completion date is added
                    if request.form.get('completionDate'):
                        # Move record to completion_data.xlsx
                         # Copy selected columns to completion_data.xlsx
                        completion_workbook = load_workbook('completion_data.xlsx')
                        completion_sheet = completion_workbook.active

                        # Find the first empty row
                        completion_row_index = 2
                        while completion_sheet.cell(row=completion_row_index, column=1).value is not None:
                            completion_row_index += 1

                        # Assign specific values to each column in completion_data.xlsx
                        completion_sheet.cell(row=completion_row_index, column=1, value=batch_sheet.cell(row=row_index, column=1).value)
                        completion_sheet.cell(row=completion_row_index, column=2, value=batch_sheet.cell(row=row_index, column=2).value)
                        completion_sheet.cell(row=completion_row_index, column=3, value=batch_name)
                        completion_sheet.cell(row=completion_row_index, column=4, value=batch_sheet.cell(row=row_index, column=5).value)
                        completion_sheet.cell(row=completion_row_index, column=5, value=batch_sheet.cell(row=row_index, column=3).value)
                        completion_sheet.cell(row=completion_row_index, column=6, value=request.form.get('completionDate'))
                        completion_sheet.cell(row=completion_row_index, column=7, value='')
                        completion_sheet.cell(row=completion_row_index, column=8, value='')
                        completion_sheet.cell(row=completion_row_index, column=9, value='')
                        completion_sheet.cell(row=completion_row_index, column=10, value='')
                        completion_sheet.cell(row=completion_row_index, column=11, value=batch_sheet.cell(row=row_index, column=8).value)

                        batch_sheet.delete_rows(row_index)
                        batch_workbook.save('batch_data.xlsx')
                        completion_workbook.save('completion_data.xlsx')


                    response_data = {'status': 'success', 'message': 'Update Successful'}
                    return jsonify(response_data)

    except Exception as e:
        print(f"Error: {e}")

    response_data = {'status': 'error', 'message': 'Update Failed'}
    return jsonify(response_data)

# Add this route in your Flask application
@app.route('/update_completion_data', methods=['POST'])
def update_completion_data():
    try:
        student_id = request.form.get('studentId')
        completion_date = request.form.get('completionDate')
        exam_date = request.form.get('examDate')
        certificate_number = request.form.get('certificateNumber')
        issue_certificate_date = request.form.get('issueCertificateDate')
        receiver_name = request.form.get('receiverName')
        final_fees = request.form.get('finalFees')

        print(f"Received request to update completion data for Student ID {student_id}")

        # Load completion workbook
        completion_workbook = load_workbook('completion_data.xlsx')
        completion_sheet = completion_workbook.active

        # Find the row with the matching student ID
        for row in completion_sheet.iter_rows(min_row=2, max_col=12, max_row=completion_sheet.max_row):
            if row[0].value == int(student_id):
                # Update the completion data
                row[6].value = exam_date
                row[7].value = certificate_number
                row[8].value = issue_certificate_date
                row[9].value = receiver_name

                completion_workbook.save('completion_data.xlsx')

                response_data = {'status': 'success', 'message': 'Update Successful'}
                return jsonify(response_data)

    except Exception as e:
        print(f"Error: {e}")

    response_data = {'status': 'error', 'message': 'Update Failed'}
    return jsonify(response_data)

@app.route('/add_inquiry', methods=['POST'])
def add_inquiry():
    ensure_inquiry_file()
    try:
        inquiry_workbook = load_workbook('inquiry_data.xlsx')
        inquiry_sheet = inquiry_workbook['Sheet']

        inquiry_date = request.form.get('inquiryDate')
        city = request.form.get('city')
        mono = request.form.get('moNo')
        name = request.form.get('name')
        course = request.form.get('course')
        batch = request.form.get('batch')
        start_date = request.form.get('startDate')

        # Find the first empty row
        row_index = 2
        while inquiry_sheet.cell(row=row_index, column=1).value is not None:
            row_index += 1

        # Use the same ID for both student and batch data
        student_id = row_index - 1

        # Update the student data Excel file with the new data
        inquiry_sheet.cell(row=row_index, column=1, value=student_id)
        inquiry_sheet.cell(row=row_index, column=2, value=inquiry_date)
        inquiry_sheet.cell(row=row_index, column=3, value=name)
        inquiry_sheet.cell(row=row_index, column=4, value=city)
        inquiry_sheet.cell(row=row_index, column=5, value=mono)
        inquiry_sheet.cell(row=row_index, column=6, value=course)
        inquiry_sheet.cell(row=row_index, column=7, value=batch)
        inquiry_sheet.cell(row=row_index, column=8, value=start_date)

        inquiry_workbook.save('inquiry_data.xlsx')

        response_data = {'status': 'success', 'message': 'Update Successful'}
        return jsonify(response_data)

    except Exception as e:
        print(f"Error: {e}")

    response_data = {'status': 'error', 'message': 'Update Failed'}
    return jsonify(response_data)

@app.route('/get_student_list')
def get_student_list():
    try:
        batch_workbook = load_workbook('batch_data.xlsx')
        batch_sheets = batch_workbook.sheetnames

        student_list = []

        for batch_name in batch_sheets:
            batch_sheet = batch_workbook[batch_name]

            col_index = None
            for cell in batch_sheet[1]:
                if cell.value == 'ID':
                    col_index = cell.column

            if col_index is not None:
                for row in batch_sheet.iter_rows(min_row=2, max_col=1, max_row=batch_sheet.max_row):
                    student_id = row[0].value
                    student_name = batch_sheet.cell(row=row[0].row, column=2).value
                    student_list.append({'id': student_id, 'name': student_name})

        return jsonify({'students': student_list})

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': 'Failed to fetch student list'})
    

@app.route('/get_completion_student_list')
def get_completion_student_list():
    try:
        completion_workbook = load_workbook('completion_data.xlsx',read_only=True)
        completion_sheet = completion_workbook.active

        student_list = []

        for row in completion_sheet.iter_rows(min_row=2, max_col=2, max_row=completion_sheet.max_row):
            student_id = row[0].value
            student_name = row[1].value
            student_list.append({'id': student_id, 'name': student_name})

        return jsonify({'students': student_list})

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': 'Failed to fetch student list for completion data'})

@app.route('/get_current_registration_number')
def get_current_registration_number():
    try:
        # Load the student workbook
        student_workbook = load_workbook('student_data.xlsx',read_only=True)
        student_sheet = student_workbook['Sheet']

        # Find the first empty row
        row_index = 2
        while student_sheet.cell(row=row_index, column=1).value is not None:
            row_index += 1

        # Use the same ID for both student and batch data
        registration_number = row_index - 1
        # Return the current registration number as JSON
        return jsonify({'registrationNumber': registration_number})
    except Exception as e:
        return jsonify({'error': str(e)})
    
@app.route('/get_all_past_students', methods=['GET'])
def get_all_past_students():
    try:
        workbook = load_workbook('student_data.xlsx',read_only=True)
        sheet = workbook['Sheet']

        # Fetch details of all past students
        past_students = [{'id': row[0].value, 'name': row[1].value} for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row)]

        workbook.close()
        return jsonify(past_students)

    except Exception as e:
        print(f"Error: {e}")

    workbook.close()
    return jsonify({'error': 'Failed to fetch past students'})

# Route to fetch details of a specific past student based on ID
@app.route('/fetch_past_student_details', methods=['GET'])
def fetch_past_student_details():
    student_id = request.args.get('id')

    try:
        workbook = load_workbook('student_data.xlsx',read_only=True)
        sheet = workbook['Sheet']

        # Find the row with the given student ID
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            if row[0].value == int(student_id):  # Assuming ID is in the first column
                student_details = {
                    'name': row[1].value,
                    'address': row[2].value,
                    'city': row[3].value,
                    'moNo1': row[4].value,
                    'moNo2': row[5].value,
                    'standard': row[6].value,
                    'school':row[7].value,
                    'dob':row[8].value,
                    'occ':row[9].value,
                    'course1': row[10].value,
                    'courseWhere': row[11].value,
                    'courseWhich': row[12].value,
                }
                workbook.close()
                return jsonify(student_details)

    except Exception as e:
        print(f"Error: {e}")
    
    workbook.close()
    return jsonify({'error': 'Student details not found'})

@app.route('/')
def index():
    return render_template('layout.html')

@app.route('/index')
def initial():
    return render_template('index.html')

@app.route('/update_excel', methods=['POST'])
def update_excel():
    try:
        data = request.form.to_dict()
        text=save_to_excel(data)
        if text=="Admission Successful":
            response_data = {'status': 'success', 'message': 'Update Successful'}
            return jsonify(response_data)
        else:
            response_data = {'status': 'error', 'message': 'Batch is full. Maximum admissions reached.'}
            return jsonify(response_data)
    except Exception as e:
        response_data = {'status': 'error', 'message': str(e)}
        return jsonify(response_data)

@app.route('/add_individual_data')
def add_individual_data():
    return render_template('add_individual_data.html')

@app.route('/add_completion_data')  
def add_completion_data(): 
    return render_template('add_completion_data.html')

@app.route('/add_inquiry_data')  
def add_inquiry_data():  
    return render_template('add_inquiry_data.html')

@app.route('/edit_details')  
def edit_details():  
    return render_template('edit_batch.html')

@app.route('/display_batch')
def display_batch():
    return render_template('display_batch.html')

@app.route('/get_student_details_from_completion')
def get_student_details_from_completion():
    student_id = request.args.get('id')
    try:
        # Load completion data workbook
        completion_workbook = load_workbook('completion_data.xlsx')
        completion_sheet = completion_workbook.active

        # Initialize an empty dictionary to store student details
        student_details = {}

        # Find the row with the provided student ID
        for row in completion_sheet.iter_rows(min_row=2, max_col=16, max_row=completion_sheet.max_row):
            if row[0].value == int(student_id):
                # Populate the student_details dictionary with column names and values
                column_names = ['ID', 'Name', 'Batch','Course','Mobile No','Completion Date', 'Exam Date', 'Certificate Number', 'Issue Certificate Date', 'Receiver Name', 'Final Fees']
                
                for col_index in range(len(column_names)):
                    student_details[column_names[col_index]] = row[col_index].value

                break
        print(student_details)

    except Exception as e:
        print(f"Error: {e}")
    return jsonify(student_details)
    
@app.route('/add_individual_data', methods=['POST'])
def add_individual_data_route():
    try:
        data = request.form.to_dict()

        # Add PC number information to data
        data['pcNumber'] = int(request.form.get('pcNumber'))

        update_remaining_data(data)
        response_data = {'status': 'success', 'message': 'Update Successful'}
        return jsonify(response_data)
    except Exception as e:
        response_data = {'status': 'error', 'message': str(e)}
        return jsonify(response_data)

@app.route('/get_student_details')
def get_student_details():
    student_id = request.args.get('id')

    student_details = {}
    try:
        batch_workbook = load_workbook('batch_data.xlsx')
        batch_sheets = batch_workbook.sheetnames


        for batch_name in batch_sheets:
            batch_sheet = batch_workbook[batch_name]

            col_index = None
            for cell in batch_sheet[1]:
                if cell.value == 'ID':
                    col_index = cell.column

            if col_index is not None:
                row_index = None
                for row in batch_sheet.iter_rows(min_row=2, max_col=1, max_row=batch_sheet.max_row):
                    if row[0].value == int(student_id):
                        row_index = row[0].row
                        break
                if row_index is not None:
                    student_details = {
                        'batch': batch_sheet.title,
                        'name': batch_sheet.cell(row=row_index, column=2).value,
                        'course': batch_sheet.cell(row=row_index, column=5).value,
                        'completionDate': batch_sheet.cell(row=row_index, column=16).value,
                        'installmentDate1': batch_sheet.cell(row=row_index, column=10).value,
                        'installmentAmount1': batch_sheet.cell(row=row_index, column=9).value,
                        'installmentDate2': batch_sheet.cell(row=row_index, column=12).value,
                        'installmentAmount2': batch_sheet.cell(row=row_index, column=11).value,
                        'installmentDate3': batch_sheet.cell(row=row_index, column=14).value,
                        'installmentAmount3': batch_sheet.cell(row=row_index, column=13).value,
                        'amounttobePaid': batch_sheet.cell(row=row_index, column=15).value,
                        'remainingFees': batch_sheet.cell(row=row_index, column=12).value,
                        'pcNumber': batch_sheet.cell(row=row_index, column=6).value,
                        'finalFees': batch_sheet.cell(row=row_index, column=8).value,
                    }
                    break

    except Exception as e:
        print(f"Error: {e}")

    return jsonify(student_details)

def copy_and_move_record(student_id, old_batch, new_batch, new_pc_number):
    try:
        # Load batch workbook
        batch_workbook = load_workbook('batch_data.xlsx')

        # Find the old batch sheet and new batch sheet
        old_batch_sheet = batch_workbook[old_batch]
        new_batch_sheet = batch_workbook[new_batch]

        # Find the row corresponding to the student ID in the old batch
        row_index = None
        for row in old_batch_sheet.iter_rows(min_row=2, max_col=1, max_row=old_batch_sheet.max_row):
            if row[0].value == int(student_id):
                row_index = row[0].row
                break

        if row_index is not None:
            print(f"Found Student ID {student_id} in Batch {old_batch} at Row {row_index}")

            # Copy the record to a variable
            record = []
            for cell in old_batch_sheet[row_index]:
                record.append(cell.value)

            # Delete the record from the old batch
            old_batch_sheet.delete_rows(row_index)

            # Add the record to the new batch with updated information
            new_batch_sheet.append(record[0:5]+[new_pc_number] + record[6:])  # Adjust column indices as needed

            # Save changes to batch workbook
            batch_workbook.save('batch_data.xlsx')

            return True

    except Exception as e:
        print(f"Error: {e}")

    return False

@app.route('/get_students_in_batch', methods=['GET'])
def get_students_in_batch():
    try:
        batch_name = request.args.get('batch')
        batch_workbook = load_workbook('batch_data.xlsx')
        batch_sheet = batch_workbook[batch_name]

        student_list = []
        col_index = None
        for cell in batch_sheet[1]:
            if cell.value == 'ID':
                col_index = cell.column

        if col_index is not None:
            for row in batch_sheet.iter_rows(min_row=2, max_col=16, max_row=batch_sheet.max_row):
                student_id = row[0].value
                student_name = row[1].value
                student_mono = row[2].value
                student_city = row[3].value
                student_course = row[4].value
                student_pcno = row[5].value
                student_feesrem = row[14].value
                student_list.append({'id': student_id, 'name': student_name, 'mono': student_mono, 'city':student_city, 'course':student_course,'pcno':student_pcno,'feesrem':student_feesrem})
        
        print(student_list)
        return jsonify({'students': student_list})
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': str(e)})
    
@app.route('/change_batch_pc', methods=['POST'])
def change_batch_pc():
    try:
        student_id = request.form.get('studentId')
        old_batch = request.form.get('oldBatch')
        old_pc_number = request.form.get('oldPCNumber')
        new_batch = request.form.get('newBatch')
        new_pc_number = request.form.get('pcNumber')

        print(f"Received request to change batch and PC for Student ID {student_id}")

        batch_workbook = load_workbook('batch_data.xlsx')
        new_batch_sheet = batch_workbook[new_batch]
        if new_batch_sheet.max_row >22:
            return jsonify({'status': 'error', 'message': 'New batch has reached the maximum capacity of 21 students.'})


        # Copy and move the record to the new batch
        if copy_and_move_record(student_id, old_batch, new_batch, new_pc_number):
            student_workbook = load_workbook('student_data.xlsx')
            student_sheet = student_workbook.active

            for row in student_sheet.iter_rows(min_row=2, max_col=1, max_row=student_sheet.max_row):
                if row[0].value == int(student_id):
                    student_sheet.cell(row=row[0].row, column=16, value=new_batch)
                    break

            student_workbook.save('student_data.xlsx')

            return jsonify({'status': 'success', 'message': 'Batch and PC updated successfully'})

    except Exception as e:
        print(f"Error: {e}")

    return jsonify({'status': 'error', 'message': 'Failed to update batch and PC'})

# Add this function to get unassigned PCs
@app.route('/get_unassigned_pcs')
def get_unassigned_pcs_route():
    try:
        batch_name = request.args.get('batch')
        available_pcs = get_unassigned_pcs(batch_name)
        return jsonify({'pcs': available_pcs})
    except Exception as e:
        return jsonify({'error': str(e)})

# Add this function to get unassigned PCs
def get_unassigned_pcs(batch_name):
    assigned_pcs = get_assigned_pcs(batch_name)
    max_pcs = 22
    all_pcs = set(range(1, max_pcs + 1))
    unassigned_pcs = list(all_pcs - assigned_pcs)
    print(unassigned_pcs)
    return unassigned_pcs

def get_assigned_pcs(batch_name):
    assigned_pcs = set()

    batch_workbook = load_workbook('batch_data.xlsx')

    if batch_name in batch_workbook.sheetnames:
        batch_sheet = batch_workbook[batch_name]

        for row in batch_sheet.iter_rows(min_row=2, max_col=1, max_row=batch_sheet.max_row):
            pc_number = batch_sheet.cell(row=row[0].row, column=6).value
            if pc_number is not None:
                assigned_pcs.add(int(pc_number))
    print(assigned_pcs)
    return assigned_pcs

if __name__ == '__main__':
    app.run(debug=True)
