# app.py
from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
from pathlib import Path
import openpyxl

app = Flask(__name__)

def ensure_excel_file():
    file_path = 'data.xlsx'
    if not Path(file_path).is_file():
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['ID', 'Name', 'Address', 'City', 'Course Done','Course Name', 'Mobile No 1', 'Mobile No 2',
                      'Date of Birth', 'Standard Studying', 'Course', 'Starting Date', 'Batch',
                      'Fees', 'Discount', 'Final Fees'])
        workbook.save(file_path)

def save_to_excel(data):
    ensure_excel_file()

    workbook = load_workbook('data.xlsx')
    sheet = workbook['Sheet']

    # Find the first empty row
    row_index = 2
    while sheet.cell(row=row_index, column=1).value is not None:
        row_index += 1

    # Update the Excel file with the new data
    sheet.cell(row=row_index, column=1, value=row_index - 1)
    sheet.cell(row=row_index, column=2, value=data['name'])
    sheet.cell(row=row_index, column=3, value=data['address'])
    sheet.cell(row=row_index, column=4, value=data['city'])
    sheet.cell(row=row_index, column=5, value=data['course1'])
    sheet.cell(row=row_index, column=6, value=data['courseName'])
    sheet.cell(row=row_index, column=7, value=data['moNo1'])
    sheet.cell(row=row_index, column=8, value=data['moNo2'])
    sheet.cell(row=row_index, column=9, value=data['dob'])
    sheet.cell(row=row_index, column=10, value=data['standard'])
    sheet.cell(row=row_index, column=11, value=data['course'])
    sheet.cell(row=row_index, column=12, value=data['startDate'])
    sheet.cell(row=row_index, column=13, value=data['batch'])
    sheet.cell(row=row_index, column=14, value=data['fees'])
    sheet.cell(row=row_index, column=15, value=data['discount'])
    sheet.cell(row=row_index, column=16, value=data['finalFees'])

    workbook.save('data.xlsx')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/update_excel', methods=['POST'])
def update_excel():
    try:
        data = request.form.to_dict()
        save_to_excel(data)
        return jsonify({'message': 'Data updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)})

if __name__ == '__main__':
    app.run(debug=True)
