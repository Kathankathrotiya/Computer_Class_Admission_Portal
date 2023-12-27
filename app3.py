# app.py
from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from flask_cors import CORS
app = Flask(__name__)
CORS(app)

def ensure_excel_files():
    ensure_student_file()
    ensure_batch_file()
    ensure_completion_file()

def ensure_student_file():
    file_path = 'student_data.xlsx'
    if not Path(file_path).is_file():
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['ID', 'Name', 'Batch', 'Completion Date', 'Exam Date', 'Certificate Date', 'Issue Certificate Date', 'Receiver Name', 'Fees Paid', 'Remaining Fees', 'PC Number'])
        workbook.save(file_path)

def ensure_batch_file():
    file_path = 'batch_data.xlsx'
    if not Path(file_path).is_file():
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

def ensure_completion_file():
    file_path = 'completion_data.xlsx'
    if not Path(file_path).is_file():
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['ID', 'Name', 'Batch', 'Completion Date', 'Exam Date', 'Certificate Date', 'Issue Certificate Date', 'Receiver Name', 'Fees Paid', 'Remaining Fees', 'PC Number'])
        workbook.save(file_path)

def save_to_excel(data):
    ensure_excel_files()

    # Save data to student_data.xlsx
    student_workbook = load_workbook('student_data.xlsx')
    student_sheet = student_workbook['Sheet']

    # Find the first empty row
    row_index = 2
    while student_sheet.cell(row=row_index, column=1).value is not None:
        row_index += 1

    # Use the same ID for both student and batch data
    student_id = row_index - 1

    # Update the student data Excel file with the new data
    student_sheet.cell(row=row_index, column=1, value=student_id)
    student_sheet.cell(row=row_index, column=2, value=data['name'])
    student_sheet.cell(row=row_index, column=3, value=data['address'])
    student_sheet.cell(row=row_index, column=4, value=data['city'])
    student_sheet.cell(row=row_index, column=5, value=data['course1'])
    student_sheet.cell(row=row_index, column=6, value=data['courseName'])
    student_sheet.cell(row=row_index, column=7, value=data['moNo1'])
    student_sheet.cell(row=row_index, column=8, value=data['moNo2'])
    student_sheet.cell(row=row_index, column=9, value=data['dob'])
    student_sheet.cell(row=row_index, column=10, value=data['standard'])
    student_sheet.cell(row=row_index, column=11, value=data['course'])
    student_sheet.cell(row=row_index, column=12, value=data['startDate'])
    student_sheet.cell(row=row_index, column=13, value=data['batch'])
    student_sheet.cell(row=row_index, column=14, value=data['fees'])
    student_sheet.cell(row=row_index, column=15, value=data['discount'])
    student_sheet.cell(row=row_index, column=16, value=data['finalFees'])

    student_workbook.save('student_data.xlsx')

    # Save data to batch_data.xlsx
    batch_workbook = load_workbook('batch_data.xlsx')

    batch_name = data['batch']
    if batch_name not in batch_workbook.sheetnames:
        batch_workbook.create_sheet(title=batch_name)

    batch_sheet = batch_workbook[batch_name]

    # Add header row if sheet is newly created
    if batch_sheet.max_row == 1:
        batch_sheet.append(['ID', 'Name', 'Mobile No 1', 'City','Course', 'Completion Date', 'Exam Date',
                            'Certificate Date', 'Issue Certificate Date', 'Receiver Name', 'Fees Paid',
                            'Remaining Fees', 'PC Number'])

    # Find the first empty row
    row_index = batch_sheet.max_row + 1

    # Set the initial remaining fees to the final fees submitted in the form
    initial_remaining_fees = data['finalFees']

    # Update the batch data Excel file with the new data
    batch_sheet.cell(row=row_index, column=1, value=student_id)
    batch_sheet.cell(row=row_index, column=2, value=data['name'])
    batch_sheet.cell(row=row_index, column=3, value=data['moNo1'])
    batch_sheet.cell(row=row_index, column=4, value=data['city'])
    batch_sheet.cell(row=row_index, column=5, value=data['course'])
    batch_sheet.cell(row=row_index, column=6, value='')
    batch_sheet.cell(row=row_index, column=7, value='')
    batch_sheet.cell(row=row_index, column=8, value='')
    batch_sheet.cell(row=row_index, column=9, value='')
    batch_sheet.cell(row=row_index, column=10, value='')
    batch_sheet.cell(row=row_index, column=11, value='')
    batch_sheet.cell(row=row_index, column=12, value=initial_remaining_fees)

    batch_workbook.save('batch_data.xlsx')

@app.route('/update_remaining_data', methods=['POST'])
def update_remaining_data():
    try:
        student_id = request.form.get('studentId')
        batch_name = request.form.get('batch')

        print(f"Received request to update data for Student ID {student_id} in Batch {batch_name}")

        batch_workbook = load_workbook('batch_data.xlsx')

        if batch_name in batch_workbook.sheetnames:
            batch_sheet = batch_workbook[batch_name]

            col_index = None
            for cell in batch_sheet[1]:
                if cell.value == 'ID':
                    col_index = cell.column

            if col_index is not None:
                row_index = None
                for row in batch_sheet.iter_rows(min_row=2, max_col=1, max_row=batch_sheet.max_row):
                    if row[0].value == int(student_id):
                        row_index = row[0].row
                        break

                if row_index is not None:
                    print(f"Found Student ID {student_id} in Batch {batch_name} at Row {row_index}")

                    # Update remaining data based on the submitted form data
                    batch_sheet.cell(row=row_index, column=6).value = request.form.get('completionDate')
                    batch_sheet.cell(row=row_index, column=7).value = request.form.get('examDate')
                    batch_sheet.cell(row=row_index, column=8).value = request.form.get('certificateDate')
                    batch_sheet.cell(row=row_index, column=9).value = request.form.get('issueCertificateDate')
                    batch_sheet.cell(row=row_index, column=10).value = request.form.get('receiverName')
                    batch_sheet.cell(row=row_index, column=11).value = request.form.get('feesPaid')
                    batch_sheet.cell(row=row_index, column=12).value = request.form.get('remainingFees')
                    batch_sheet.cell(row=row_index, column=13).value = request.form.get('pcNumber', '')  # Ensure pcNumber is correctly handled

                    batch_workbook.save('batch_data.xlsx')

                    # Check if completion date is added
                    if request.form.get('completionDate'):
                        # Move record to completion_data.xlsx
                        completion_workbook = load_workbook('completion_data.xlsx')
                        completion_sheet = completion_workbook.active

                        # Find the first empty row
                        completion_row_index = 2
                        while completion_sheet.cell(row=completion_row_index, column=1).value is not None:
                            completion_row_index += 1

                        # Move the record from batch_data.xlsx to completion_data.xlsx
                        for col in range(1, 12):
                            value = batch_sheet.cell(row=row_index, column=col).value
                            completion_sheet.cell(row=completion_row_index, column=col, value=value)

                        batch_sheet.delete_rows(row_index)
                        batch_workbook.save('batch_data.xlsx')
                        completion_workbook.save('completion_data.xlsx')

                    return "Update Successful"

    except Exception as e:
        print(f"Error: {e}")

    return "Update Failed"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/update_excel', methods=['POST'])
def update_excel():
    try:
        data = request.form.to_dict()
        save_to_excel(data)
        return jsonify({'message': 'Data updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/add_individual_data')
def add_individual_data():
    # batch_name = request.args.get('batch')
    # available_pcs = get_available_pcs(batch_name)
    return render_template('add_individual_data.html')



# @app.route('/add_completion_data')  # Changed route name
# def add_completion_data():  # Changed route function name
#     return render_template('add_completion_data.html')

@app.route('/add_individual_data', methods=['POST'])
def add_individual_data_route():
    try:
        data = request.form.to_dict()

        # Add PC number information to data
        data['pcNumber'] = int(request.form.get('pcNumber'))

        update_remaining_data(data)
        return jsonify({'message': 'Individual data added successfully'})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/get_student_details')
def get_student_details():
    student_id = request.args.get('id')

    try:
        batch_workbook = load_workbook('batch_data.xlsx')
        batch_sheets = batch_workbook.sheetnames

        student_details = {}

        for batch_name in batch_sheets:
            batch_sheet = batch_workbook[batch_name]

            col_index = None
            for cell in batch_sheet[1]:
                if cell.value == 'ID':
                    col_index = cell.column

            if col_index is not None:
                row_index = None
                for row in batch_sheet.iter_rows(min_row=2, max_col=1, max_row=batch_sheet.max_row):
                    if row[0].value == int(student_id):
                        row_index = row[0].row
                        break
                if row_index is not None:
                    student_details = {
                        'batch': batch_sheet.title,
                        'name': batch_sheet.cell(row=row_index, column=2).value,
                        'course': batch_sheet.cell(row=row_index, column=5).value,
                        'completionDate': batch_sheet.cell(row=row_index, column=6).value,
                        'examDate': batch_sheet.cell(row=row_index, column=7).value,
                        'certificateDate': batch_sheet.cell(row=row_index, column=8).value,
                        'issueCertificateDate': batch_sheet.cell(row=row_index, column=9).value,
                        'receiverName': batch_sheet.cell(row=row_index, column=10).value,
                        'feesPaid': batch_sheet.cell(row=row_index, column=11).value,
                        'remainingFees': batch_sheet.cell(row=row_index, column=12).value,
                        'pcNumber': batch_sheet.cell(row=row_index, column=13).value,
                    }
                    break

    except Exception as e:
        print(f"Error: {e}")

    return jsonify(student_details)

# Add this function to get unassigned PCs
@app.route('/get_unassigned_pcs')
def get_unassigned_pcs_route():
    try:
        batch_name = request.args.get('batch')
        available_pcs = get_unassigned_pcs(batch_name)
        return jsonify({'pcs': available_pcs})
    except Exception as e:
        return jsonify({'error': str(e)})

# Add this function to get unassigned PCs
def get_unassigned_pcs(batch_name):
    assigned_pcs = get_assigned_pcs(batch_name)
    max_pcs = 22
    all_pcs = set(range(1, max_pcs + 1))
    unassigned_pcs = list(all_pcs - assigned_pcs)
    print(unassigned_pcs)
    return unassigned_pcs

def get_assigned_pcs(batch_name):
    assigned_pcs = set()

    batch_workbook = load_workbook('batch_data.xlsx')

    if batch_name in batch_workbook.sheetnames:
        batch_sheet = batch_workbook[batch_name]

        for row in batch_sheet.iter_rows(min_row=2, max_col=1, max_row=batch_sheet.max_row):
            pc_number = batch_sheet.cell(row=row[0].row, column=13).value
            if pc_number is not None:
                assigned_pcs.add(int(pc_number))
    print(assigned_pcs)
    return assigned_pcs

if __name__ == '__main__':
    app.run(debug=True)
